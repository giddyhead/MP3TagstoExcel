import mutagen, xlrd, glob, re, openpyxl, os, pygal
# from mutagen.easyid3 import EasyID3
from os import walk
from pprint import pprint
from tinytag import TinyTag, TinyTagException
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from mp3_tagger import MP3File
from string import ascii_uppercase
from mutagen.mp3 import MP3
from openpyxl.workbook import Workbook

tracks = []
gettags = []

def ExtractMP3TagtoExcel():

        for root, dirs, files, in os.walk(r'C:\Users\mrdrj\Desktop\sdf'):
            for name in files:
                if name.endswith(('.mp3', '.m4a', '.flac', '.alac')):
                    tracks.append(name)  # Add Media Files
            try:
                track_filepath = os.path.join(root, name)
                temp_track = TinyTag.get(track_filepath)
                mp3 = MP3File(track_filepath)
            except TinyTagException as err:
                    print(err)
                    continue

            gettags2 =[temp_track.album, temp_track.albumartist, temp_track.artist, temp_track.audio_offset,
                temp_track.bitrate, temp_track.comment, temp_track.composer, temp_track.disc,
                temp_track.disc_total, temp_track.duration, temp_track.filesize, temp_track.genre,
                temp_track.samplerate, temp_track.title, temp_track.track, temp_track.track_total,
                temp_track.year]  # Add Tags to list

            for x in range(len(gettags2)):
                # append slice of gettags2, containing the entire gettags2
                gettags.append(gettags2[:])

            #os.path.join(root, name)
            header = ['album', 'albumartist', 'artist', 'audio_offset', 'bitrate', 'comment', 'composer', 'disc',
          'disc_total', 'duration', 'filesize', 'genre', 'samplerate', 'title', 'track', 'track_total', 'year']

            wb = Workbook()
            new_data = gettags
            dest_filename = '11empty_book11.xlsx'
            ws1 = wb.active
            ws1.title = "MP3 Tags"
            ws2 = wb.create_sheet(title="Set")
            ws1.append(header[:])

            tags = []
            for row in new_data:  # Number of Rows
             # tags.append(new_data[:]) #Add to Tag List
                row = set(row)
                ws1.append(row)
                tags.append(tuple(row))
            print(row)

            wb.save(filename=dest_filename)

ExtractMP3TagtoExcel()

